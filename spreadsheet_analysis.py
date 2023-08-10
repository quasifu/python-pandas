import openpyxl

def parse_excel_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet_info = {}

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        columns = []

        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col_idx).value
            columns.append(cell_value)

        worksheet_info[sheet_name] = columns

    return worksheet_info

if __name__ == "__main__":
    excel_file_path = "SimplySails.xlsx"

    try:
        worksheet_info = parse_excel_file(excel_file_path)

        for sheet_name, columns in worksheet_info.items():
            print(f"Worksheet: {sheet_name}")
            print("Columns:", ", ".join(str(col) for col in columns))
            print("-" * 30)

    except FileNotFoundError:
        print(f"File not found: {excel_file_path}")
    except Exception as e:
        print(f"Error occurred: {str(e)}")
